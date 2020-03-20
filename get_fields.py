# -*- coding: UTF-8 -*-
'''
@Author : mrchi
@Time   : 2020/3/16 18:12
@Project: Operation_Excel
@Website: https://www.mrchi.cn
@File   : get_fields.py
@Desc   :
@Version: 2.1
'''

import time

import openpyxl as excel
import os
import random
from fanyi.request_youdao import get_request_youdao
import warnings
from pinyin.convert_char import Convert_to_char
#  忽略 警告信息
warnings.filterwarnings("ignore")


class Operation_excel_file():

    def __init__(self, base_dir):
        self.convert_type = None
        self.base_dir = base_dir + "\\"
        self.table_name = input(
            "*\t\t\tPlease enter generated table  name" + "\t" * 4 + "  *\n*" + "\t" * 15 + "  *\n*  " + "* " * 5 + ":")
        self.file_path = ""
        self.sql_prefix = "DROP TABLE IF EXISTS `%s` ;\n CREATE TABLE `%s` ( " % (self.table_name, self.table_name)
        self.hive_sql_postfix = ") COMMENT \"%s\" \nROW FORMAT DELIMITED FIELDS TERMINATED BY '\\t'\n STORED AS ORCFILE;" % self.table_name
        self.mysql_sql_postfix = ") COMMENT \"%s\"  ENGINE = InnoDB AUTO_INCREMENT = 351 CHARACTER SET = utf8mb4 COLLATE = utf8mb4_bin ROW_FORMAT = Compact; " % self.table_name

        self.ddl_file_name = ""
        self.src_fields = []
        self.translate_result = []
        self.translate_column_name()

    # Getting  path from excel file
    def dynamic_get_excel_by_path(self):
        file_list = os.listdir(self.base_dir)
        if len(file_list):
            for x in range(len(file_list)):
                if str(file_list[x]).find("xlsx") >= 0:
                    self.ddl_file_name = file_list[x].split(".")[0] + "__ddl.sql"
                    return file_list[x]

    # Getting column name by openpyxl module
    def get_excel_by_column_name(self):
        get_excel_by_path = self.dynamic_get_excel_by_path()
        if get_excel_by_path:
            work_book = excel.load_workbook(self.base_dir + get_excel_by_path)
            # hard code
            # work_sheets = work_book["Worksheet"]

            # Getting all sheet name from work_book object
            work_sheet_name = work_book.sheetnames

            print("\n\n\n  \t\t\t\t---->>   All sheet %s \n\n" % work_sheet_name)

            # todo 迭代后添加用户选择操作，让用户选择使用哪个sheet

            # Getting  frist  sheet
            first_sheet = work_book[work_sheet_name[0]]

            # Getting all fields name
            for i in range(1, first_sheet.max_column + 1):
                cell_value = first_sheet.cell(row=1, column=i).value
                if cell_value is not None and len(cell_value) > 0:
                    # If value not None of append to list
                    self.src_fields.append(cell_value)
            print("\t\t\t\t---->>   Total of %s column\n\n" % len(self.src_fields))
            print("\t\t\t\t---->>   File Name%s \n\n" % self.ddl_file_name)

            return self.src_fields
        else:
            print("Unable to get Excel file !")
            exit(0)

    # Transfor column name by youdao
    def translate_column_name(self):
        chinese_char_list = self.get_excel_by_column_name()
        chinese_char_list_count = chinese_char_list.copy()
        if chinese_char_list:
            for chinese in chinese_char_list:
                chinese_char_list_count.remove(chinese)
                # print(len(chinese))
                # print(type(chinese))
                # print("transformation   %s   ing ...\n\n" % chinese)

                if len(chinese) > 0 and chinese is not None:
                    while True:
                        if self.convert_type is None:
                            self.convert_type = input(
                                " Please select converted type: \n\n\n   \t\t\t n --->> English\n\n \t\t\t p --->>> Pinyin\n\n\n")
                        else:
                            break
                    if str(self.convert_type.lower()) == "n":
                        time.sleep(random.randint(1, 5))
                        print("[ %d/%d ] ---->>    %s   Converting ...\n\n\n" % (
                            (len(chinese_char_list) - len(chinese_char_list_count)),len(chinese_char_list), chinese))
                        result_english = get_request_youdao(chinese.strip()).start()
                        result_english = self.replace_str(result_english)
                        self.translate_result.append(result_english)
                    elif str(self.convert_type.lower()) == "p":
                        result_chars = Convert_to_char(chinese.strip()).chars
                        self.translate_result.append(result_chars)
                else:
                    print("The input is invalid , please confirm  input value \n sorry system exit !")
                    exit(0)
            print("\n" * 5 + "* " * 10 + "Show hive style " + "* " * 10 + "\n" * 3)
            return self.translate_result

    # Rinse string
    def replace_str(self, str):
        rinse_result = str.replace(" ", "_").replace("/", "_").replace(",", "").replace(
            ".",
            "").replace(
            "(",
            "").replace(
            ")", "").replace("\"", "").replace("\\", "").replace("\'", "").replace("-", "_").lower().replace("the",
                                                                                                             "").replace(
            "of", "").replace("for", "").replace("in", "").replace("at", "").replace("ings",
                                                                                     "").replace(
            "ing", "").replace("?", "_").replace("__", "_")
        # rinse_result=rinse_result.replace("","")
        return rinse_result

    # Generate file
    def generate_fields(self):
        if self.translate_result:
            hive_fields_list = []
            mysql_fields_list = []

            # Write  SQL  prefix
            if self.sql_prefix not in hive_fields_list:
                hive_fields_list.append(self.sql_prefix)
            if self.sql_prefix not in mysql_fields_list:
                mysql_fields_list.append(self.sql_prefix)

            for index in range(0, len(self.translate_result)):
                last_of_one = len(self.translate_result) - index

                # Is last one?
                if last_of_one > 1:
                    hive_fields_list.append(
                        " `%s`  STRING   COMMENT  '%s'," % (self.translate_result[index], self.src_fields[index]))

                    mysql_fields_list.append(
                        " `%s`  VARCHAR(200)   COMMENT  '%s'," % (
                            self.translate_result[index], self.src_fields[index]))
                # Last one
                else:
                    hive_fields_list.append(
                        " `%s`  STRING   COMMENT  '%s'" % (self.translate_result[index], self.src_fields[index]))

                    mysql_fields_list.append(
                        " `%s`  VARCHAR(200)   COMMENT  '%s'" % (
                            self.translate_result[index], self.src_fields[index]))

            #  生成hive与mysql的ddl

            # Add hive suffix
            if self.hive_sql_postfix not in hive_fields_list:
                hive_fields_list.append(self.hive_sql_postfix)

            # Add mysql suffix
            if self.mysql_sql_postfix not in mysql_fields_list:
                mysql_fields_list.append(self.mysql_sql_postfix)

            # Delete exist file
            if os.path.isfile(self.base_dir + self.ddl_file_name):
                os.remove(self.base_dir + self.ddl_file_name)

            # Start write hive header
            with open(self.base_dir + self.ddl_file_name, "a+") as f:
                f.writelines("-- " + "* " * 15 + " Hive  DDL" + "* " * 15 + "\n")
                f.writelines("-- " + "*" * 70 + "\n")
                f.writelines("-- 功能：\t\t\t\t 创建 什么什么 表 ( %s )" % self.table_name + "\n")
                f.writelines("-- 作者：\t\t\t\t 念迟 & https://www.mrchi.cn \n")
                f.writelines("-- 时间：\t\t\t\t %s \n" % time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
                f.writelines("-- " + "*" * 70 + "\n")
            # Iterate generate hive DDL
            for filelds in hive_fields_list:
                print(filelds)
                with open(self.base_dir + self.ddl_file_name, "a+") as f:
                    f.writelines(filelds + "\n")

            # Start write mysql  header
            with open(self.base_dir + self.ddl_file_name, "a+") as f:
                f.writelines("\n\n-- " + "* " * 15 + " MySQL  DDL " + "* " * 15 + "\n")
                f.writelines("-- " + "*" * 70 + "\n")
                f.writelines("-- 功能：\t\t\t\t 创建 什么什么 表 ( %s )" % self.table_name + "\n")
                f.writelines("-- 作者：\t\t\t\t 念迟 & https://www.mrchi.cn \n")
                f.writelines("-- 时间：\t\t\t\t %s \n" % time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
                f.writelines("-- " + "*" * 70 + "\n")

            # Iterate generate mysql ddl
            for filelds in mysql_fields_list:
                with open(self.base_dir + self.ddl_file_name, "a+") as f:
                    f.writelines(filelds + "\n")
        else:
            print("The transform content cannot be empty")
            exit(0)


if __name__ == '__main__':
    base_dir = os.path.split(os.path.realpath(__file__))[0]
    print("* " * 32 + "\n* " + "\t" * 15 + "  *" + "\n*" + "\t" * 15 + "  *")
    # print("*\n*\n* \t\t\t By: 念迟 & https://www.mrchi.cn  \n*")
    print(
        "* \t\t\t默认生成 Hive DDL 与 Mysql DDL (v2.3)" + "\t" * 3 + "  *" + "\n* " + "\t" * 15 + "  *" + "\n* " + "\t" * 15 + "  *")
    print("* " * 32 + "\n* " + "\t" * 15 + "  *" + "\n*" + "\t" * 15 + "  *")
    oef = []
    oef = Operation_excel_file(base_dir=base_dir)
    oef.generate_fields()

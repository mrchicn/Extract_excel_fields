# -*- coding: UTF-8 -*-
'''
@Author : mrchi
@Time   : 2020/3/18 13:16
@Project: Extract_excel_fields
@Website: https://www.mrchi.cn
@File   : field_process.py
@Desc   :
'''
import os

import openpyxl as excel

base_dir = os.path.split(os.path.realpath(__file__))[0] + "\\"
proccess_word_chinese = []


def replace_str(word):
    result = word.replace(" ", "_").replace("/", "_").replace(",", "").replace(
        ".",
        "").replace(
        "(",
        "").replace(
        ")", "").replace("\"", "").replace("\\", "").replace("\'", "").replace("-", "_").lower().strip().replace("the_",
                                                                                                                 "").replace(
        "_of", "").replace("for_", "").replace("in_", "").replace("nian", "year").replace("?","_")
    return result


def hive_ddl(src_fields, target_fields):
    table_name = "Hive_DDL"
    sql_suffix = "DROP TABLE IF EXISTS `%s` ;\n CREATE TABLE `%s` ( " % (table_name, table_name)
    sql_postfix = ") COMMENT `%s`\nROW FORMAT DELIMITED FIELDS TERMINATED BY ','\n STORED AS TEXTFILE;" % table_name
    for index in range(0, len(target_fields)):
        last_of_one = len(target_fields) - index
        if index == 0:
            print(sql_suffix)
            if last_of_one > 1:
                print(" `%s`  STRING   COMMENT  '%s'," % (target_fields[index], src_fields[index]))
            else:
                print(
                    " `%s`  STRING   COMMENT  '%s'" % (target_fields[index], src_fields[index]))
        else:
            print(sql_postfix)


word_chinese = []
words_english = []


# Getting  path from excel file
def dynamic_get_excel_by_path():
    file_list = os.listdir(base_dir)
    if len(file_list):
        for x in range(len(file_list)):
            if str(file_list[x]).find("xlsx") >= 0:
                return file_list[x]


# Getting column name by openpyxl module
def get_excel_by_column_name():
    get_excel_by_path = dynamic_get_excel_by_path()
    if get_excel_by_path:
        work_book = excel.load_workbook(base_dir + get_excel_by_path)
        # hard code
        # work_sheets = work_book["Worksheet"]

        # Getting all sheet name from work_book object
        work_sheet_name = work_book.sheetnames
        # print("All sheet %s " % work_sheet_name)

        # Getting  frist  sheet
        first_sheet = work_book[work_sheet_name[0]]

        # Getting all fields name
        for i in range(1, first_sheet.max_column + 1):
            cell_value = first_sheet.cell(row=1, column=i).value
            if cell_value is not None and len(cell_value) > 0:
                word_chinese.append(cell_value)
        print("Total of %s column" % str(first_sheet.max_column - 1))
        return word_chinese
    else:
        print("Unable to get Excel file !")
        exit(0)

    # for english, chinese in zip(words_english, word_chinese):
    #     english_fields = replace_str(english, )
    #     hive_ddl_result = hive_ddl(english_fields, word_chinese)


def start():
    word_chinese = get_excel_by_column_name()

    for fields in word_chinese:
        print("  STRING COMMENT  '%s'," % fields)
    print("*" * 200)
    for fields in word_chinese:
        print(fields)


def scond():
    words_english = [

"          \nThree layers of completion\n                                                      "







    ]
    for fields in words_english:
        print(replace_str(fields) + "   -cyc")
    print("*" * 200)
    for fields in words_english:
        print(replace_str(fields))


if __name__ == '__main__':
    # start()
    scond()
    # hive_ddl(proccess_word_chinese,words_english)
